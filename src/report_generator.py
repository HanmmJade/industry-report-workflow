# -*- coding: utf-8 -*-
"""
报告生成器 - 生成Markdown和Excel格式的报告

包含：
- generate_markdown_report: 生成Markdown报告
- generate_excel_template: 生成Excel模板
"""

import json
from datetime import datetime
from typing import List, Dict, Any, Optional
from pathlib import Path

# 条件导入openpyxl（可选）
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from data_models import GradingResult, GradingStats, CredibilityGrade


class ReportGenerator:
    """
    报告生成器
    
    生成标准Markdown报告和Excel模板
    """
    
    # 可信度等级颜色映射
    GRADE_COLORS = {
        "A": "🟢",
        "B": "🔵",
        "C": "🟡",
        "D": "🔴",
        "UNKNOWN": "⚪"
    }
    
    def __init__(self):
        """初始化报告生成器"""
        pass
    
    def generate_markdown_report(self, grading_results: List[GradingResult], stats: GradingStats) -> str:
        """
        生成标准Markdown报告
        
        Args:
            grading_results: 评级结果列表
            stats: 统计信息
        
        Returns:
            Markdown格式报告内容
        """
        lines = []
        
        # 1. 报告头部
        lines.append("# 行业月报数据质量报告")
        lines.append("")
        lines.append(f"> **生成时间：** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"> **数据条目数：** {stats.total_count}")
        lines.append("")
        lines.append("---")
        lines.append("")
        
        # 2. 可信度分布
        lines.append("## 一、可信度分布统计")
        lines.append("")
        lines.append("| 等级 | 数量 | 占比 | 分布 |")
        lines.append("|------|------|------|------|")
        
        grade_names = {"A": "高可信", "B": "较可信", "C": "待验证", "D": "存疑", "UNKNOWN": "未知"}
        for grade in ["A", "B", "C", "D", "UNKNOWN"]:
            count = stats.grade_distribution.get(grade, 0)
            percentage = stats.grade_percentages.get(grade, 0)
            name = grade_names.get(grade, grade)
            color = self.GRADE_COLORS.get(grade, "⚪")
            lines.append(f"| {color} {grade}级({name}) | {count} | {percentage}% | {'█' * int(percentage / 5)} |")
        
        lines.append("")
        
        # 3. 校验统计
        lines.append("## 二、校验统计")
        lines.append("")
        lines.append(f"- **校验通过率：** {round(stats.validation_pass_rate * 100, 2)}%")
        lines.append(f"- **校验通过数量：** {stats.validation_pass_count}/{stats.total_count}")
        lines.append("")
        
        if stats.common_issues:
            lines.append("### 常见异常")
            lines.append("")
            for issue_type, count in stats.common_issues.items():
                lines.append(f"- **{issue_type}：** {count}条")
            lines.append("")
        
        if stats.common_downgrades:
            lines.append("### 常见降级原因")
            lines.append("")
            for reason, count in stats.common_downgrades.items():
                lines.append(f"- {reason}：{count}条")
            lines.append("")
        
        # 4. 详细数据
        lines.append("## 三、详细数据")
        lines.append("")
        lines.append("| 数据ID | 指标 | 最终评级 | 基础评级 | 评级理由 | 升级建议 |")
        lines.append("|--------|------|----------|----------|----------|----------|")
        
        for result in grading_results:
            grade = result.grade.value if isinstance(result.grade, CredibilityGrade) else str(result.grade)
            base_grade = result.base_grade.value if isinstance(result.base_grade, CredibilityGrade) else str(result.base_grade)
            color = self.GRADE_COLORS.get(grade, "⚪")
            
            # 截断长文本
            reason = result.grade_reason[:50] + "..." if len(result.grade_reason) > 50 else result.grade_reason
            suggestions = "; ".join(result.upgrade_suggestions[:2]) if result.upgrade_suggestions else "-"
            
            lines.append(f"| {result.data_id} | ... | {color}**{grade}级** | {base_grade}级 | {reason} | {suggestions} |")
        
        lines.append("")
        lines.append("---")
        lines.append("")
        lines.append(f"*报告生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*")
        
        return "\n".join(lines)
    
    def generate_excel_template(self, grading_results: List[GradingResult], output_path: Optional[str] = None) -> str:
        """
        生成Excel模板
        
        Args:
            grading_results: 评级结果列表
            output_path: 输出路径
        
        Returns:
            Excel文件路径
        """
        if not HAS_OPENPYXL:
            raise ImportError("需要安装openpyxl库才能生成Excel文件: pip install openpyxl")
        
        if output_path is None:
            output_dir = Path(__file__).parent.parent / "output"
            output_dir.mkdir(exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = str(output_dir / f"template_{timestamp}.xlsx")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "数据质量报告"
        
        # 样式定义
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 评级颜色
        grade_fills = {
            "A": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "B": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "C": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "D": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        }
        
        # 表头
        headers = [
            "数据ID", "指标", "最终评级", "基础评级", "评级理由",
            "交叉验证", "验证来源数", "A级验证数", "升级建议", "降级原因"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        # 数据行
        for row, result in enumerate(grading_results, 2):
            grade = result.grade.value if isinstance(result.grade, CredibilityGrade) else str(result.grade)
            base_grade = result.base_grade.value if isinstance(result.base_grade, CredibilityGrade) else str(result.base_grade)
            
            row_data = [
                result.data_id,
                "...",  # 指标（需补充）
                f"{grade}级",
                f"{base_grade}级",
                result.grade_reason,
                "是" if result.has_cross_validation else "否",
                result.validation_sources_count,
                result.validation_sources_grade_a,
                "; ".join(result.upgrade_suggestions) if result.upgrade_suggestions else "-",
                "; ".join(result.downgrade_reasons) if result.downgrade_reasons else "-"
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                
                # 评级列添加颜色
                if col == 3:  # 最终评级列
                    cell.fill = grade_fills.get(grade, PatternFill())
        
        # 调整列宽
        column_widths = [15, 30, 10, 10, 50, 10, 12, 12, 30, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        # 保存
        wb.save(output_path)
        
        return output_path
    
    def generate_summary_json(self, grading_results: List[GradingResult], stats: GradingStats) -> str:
        """
        生成JSON格式摘要
        
        Args:
            grading_results: 评级结果列表
            stats: 统计信息
        
        Returns:
            JSON字符串
        """
        summary = {
            "generated_at": datetime.now().isoformat(),
            "total_count": stats.total_count,
            "grade_distribution": stats.grade_distribution,
            "grade_percentages": stats.grade_percentages,
            "validation_pass_rate": round(stats.validation_pass_rate * 100, 2),
            "common_issues": stats.common_issues,
            "common_downgrades": stats.common_downgrades,
            "items": [result.to_dict() for result in grading_results]
        }
        
        return json.dumps(summary, ensure_ascii=False, indent=2)


# 便捷函数
def generate_markdown_report(grading_results: List[GradingResult], stats: GradingStats) -> str:
    """生成Markdown报告"""
    return ReportGenerator().generate_markdown_report(grading_results, stats)


def generate_excel_template(grading_results: List[GradingResult], output_path: Optional[str] = None) -> str:
    """生成Excel模板"""
    return ReportGenerator().generate_excel_template(grading_results, output_path)
